from decimal import Decimal
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, jsonify, current_app
from werkzeug.utils import secure_filename
from app import db, rate_limit
from sqlalchemy import func, case
from app.models import Sheep, WeightRecord, MedicalRecord, MedicalPhoto, BirthRecord, FeedRation, Pen, TreatmentTemplate, Medicine, BreedCategory, PurposeCategory, StatusCategory, Transaction, TransactionCategory, AuditLog, MatingRecord, UltrasoundRecord, UltrasoundImage, SemenInventory, QuarantineRecord, DrugInventory
from datetime import datetime, timedelta, UTC
from app.accounting_engine import AccountingEngine
import qrcode
import jdatetime
import os
import csv
import io
import xlsxwriter
import time
import random
import tempfile
from flask_login import current_user, login_required
from app.blueprints.finance import permission_required, normalize_amount_to_toman
from app.blueprints.dashboard import get_setting

livestock_bp = Blueprint('livestock', __name__)

def parse_smart_date(date_str, default=None):
    """پشتیبانی از تاریخ شمسی و میلادی در بک‌اند (مستقل از JS فرانت)"""
    if not date_str or str(date_str).strip() in ['', 'None']:
        return default
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    arabic_digits = '٠١٢٣٤٥٦٧٨٩'
    english_digits = '0123456789'
    translation_table = str.maketrans(persian_digits + arabic_digits, english_digits + english_digits)
    date_str = str(date_str).translate(translation_table).replace('/', '-').strip()
    try:
        if date_str.startswith(('13', '14')):
            p = date_str.split('-')
            return jdatetime.date(int(p[0]), int(p[1]), int(p[2])).togregorian()
        return datetime.strptime(date_str[:10], '%Y-%m-%d').date()
    except Exception:
        return default

def log_audit(action):
    try:
        user = current_user.name if current_user.is_authenticated else "سیستم/ناشناس"
        ip = request.remote_addr
        from app.models import AuditLog
        db.session.add(AuditLog(user_name=user, action=action, ip_address=ip))
        db.session.commit()
    except Exception as e:
        current_app.logger.error(f"Audit log failed: {e}")

@livestock_bp.route('/')
@login_required
@permission_required('can_view_livestock')
def index():
    today = datetime.now(UTC).date()
    maturity_days = int(get_setting('maturity_days', 240))
    maturity_date = today - timedelta(days=maturity_days)
    lambs_to_update = Sheep.query.filter(
        Sheep.gender.in_(['بره ماده', 'بره نر', 'نامشخص']),
        Sheep.birth_date <= maturity_date
    ).all()
    if lambs_to_update:
        for lamb in lambs_to_update:
            if 'ماده' in lamb.gender:
                lamb.gender = 'میش'
            else:
                lamb.gender = 'قوچ'
        db.session.commit()
    
    rations = FeedRation.query.all()
    pens = Pen.query.all()
    breeds = BreedCategory.query.all()
    statuses = StatusCategory.query.all()
    
    from sqlalchemy import func
    total_sheep = Sheep.query.filter(Sheep.is_deleted == False, Sheep.status.notin_(['تلف شده', 'مرده', 'فروخته شده'])).count()
    sick_count = Sheep.query.filter_by(status='بیمار').count()
    pregnant_count = Sheep.query.filter_by(status='آبستن').count()
    total_live_weight = db.session.query(func.sum(Sheep.weight)).filter(Sheep.status.notin_(['تلف شده', 'مرده', 'فروخته شده'])).scalar() or 0.0

    query = Sheep.query.filter(Sheep.is_deleted == False)
    search_q = request.args.get('search', '').strip()
    gender_q = request.args.get('gender', 'همه')
    breed_q = request.args.get('breed', 'همه')
    status_q = request.args.get('status', 'فعال')
    min_w = request.args.get('min_weight', type=float)
    max_w = request.args.get('max_weight', type=float)
    starred_q = request.args.get('starred')

    if search_q: query = query.filter(Sheep.ear_tag.ilike(f"%{search_q}%"))
    if gender_q != 'همه': query = query.filter(Sheep.gender == gender_q)
    if breed_q != 'همه': query = query.filter(Sheep.breed == breed_q)
    if status_q == 'فعال': query = query.filter(Sheep.status.notin_(['فروخته شده', 'تلف شده', 'مرده']))
    elif status_q == 'بایگانی': query = query.filter(Sheep.status.in_(['فروخته شده', 'تلف شده', 'مرده']))
    elif status_q != 'همه': query = query.filter(Sheep.status == status_q)
    if min_w is not None: query = query.filter(Sheep.weight >= min_w)
    if max_w is not None: query = query.filter(Sheep.weight <= max_w)
    if starred_q == '1': query = query.filter(Sheep.is_starred.is_(True))

    page = request.args.get('page', 1, type=int)
    page_size = int(get_setting('page_size', 50))
    sheeps_pagination = query.order_by(Sheep.id.desc()).paginate(page=page, per_page=page_size)

    return render_template('livestock/index.html', 
                           sheeps=sheeps_pagination, rations=rations, pens=pens, breeds=breeds, statuses=statuses,
                           total_sheep=total_sheep, sick_count=sick_count, pregnant_count=pregnant_count, total_live_weight=total_live_weight,
                           current_search=search_q, current_gender=gender_q, current_breed=breed_q, current_status=status_q, 
                           current_min_w=min_w, current_max_w=max_w, current_starred=starred_q)

@livestock_bp.route('/export')
@login_required
@permission_required('can_view_livestock')
def export_sheep():
    query = Sheep.query.filter(Sheep.is_deleted == False)
    search_q = request.args.get('search', '').strip()
    gender_q = request.args.get('gender', 'همه')
    breed_q = request.args.get('breed', 'همه')
    status_q = request.args.get('status', 'فعال')
    min_w = request.args.get('min_weight', type=float)
    max_w = request.args.get('max_weight', type=float)
    starred_q = request.args.get('starred')
    pen_id_q = request.args.get('pen_id', type=int)
    if pen_id_q: query = query.filter(Sheep.pen_id == pen_id_q)
    if search_q: query = query.filter(Sheep.ear_tag.ilike(f"%{search_q}%"))
    if gender_q != 'همه': query = query.filter(Sheep.gender == gender_q)
    if breed_q != 'همه': query = query.filter(Sheep.breed == breed_q)
    if status_q == 'فعال': query = query.filter(Sheep.status.notin_(['فروخته شده', 'تلف شده', 'مرده']))
    elif status_q == 'بایگانی': query = query.filter(Sheep.status.in_(['فروخته شده', 'تلف شده', 'مرده']))
    elif status_q != 'همه': query = query.filter(Sheep.status == status_q)
    if min_w is not None: query = query.filter(Sheep.weight >= min_w)
    if max_w is not None: query = query.filter(Sheep.weight <= max_w)
    if starred_q == '1': query = query.filter(Sheep.is_starred.is_(True))

    sheeps = query.all()

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    worksheet.right_to_left()

    headers = ['پلاک', 'نژاد', 'جنسیت', 'وزن (kg)', 'وضعیت', 'هدف', 'جیره', 'بهاربند']
    header_format = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1})
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    for row, s in enumerate(sheeps, 1):
        worksheet.write(row, 0, s.ear_tag)
        worksheet.write(row, 1, s.breed or '-')
        worksheet.write(row, 2, s.gender)
        worksheet.write(row, 3, s.weight)
        worksheet.write(row, 4, s.status)
        worksheet.write(row, 5, s.purpose or '-')
        worksheet.write(row, 6, s.ration.name if s.ration else 'ندارد')
        worksheet.write(row, 7, s.pen.name if s.pen else 'نامشخص')

    workbook.close()
    output.seek(0)

    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=livestock_export.xlsx"}
    )

@livestock_bp.route('/print')
@login_required
@permission_required('can_view_livestock')
def print_sheep():
    query = Sheep.query.filter(Sheep.is_deleted == False)
    search_q = request.args.get('search', '').strip()
    gender_q = request.args.get('gender', 'همه')
    breed_q = request.args.get('breed', 'همه')
    status_q = request.args.get('status', 'فعال')
    starred_q = request.args.get('starred')
    min_w = request.args.get('min_weight', type=float)
    max_w = request.args.get('max_weight', type=float)
    pen_id_q = request.args.get('pen_id', type=int)
    if pen_id_q: query = query.filter(Sheep.pen_id == pen_id_q)
    if search_q: query = query.filter(Sheep.ear_tag.ilike(f"%{search_q}%"))
    if gender_q != 'همه': query = query.filter(Sheep.gender == gender_q)
    if breed_q != 'همه': query = query.filter(Sheep.breed == breed_q)
    if status_q == 'فعال': query = query.filter(Sheep.status.notin_(['فروخته شده', 'تلف شده', 'مرده']))
    elif status_q == 'بایگانی': query = query.filter(Sheep.status.in_(['فروخته شده', 'تلف شده', 'مرده']))
    elif status_q != 'همه': query = query.filter(Sheep.status == status_q)
    if min_w is not None: query = query.filter(Sheep.weight >= min_w)
    if max_w is not None: query = query.filter(Sheep.weight <= max_w)
    if starred_q == '1': query = query.filter(Sheep.is_starred.is_(True))

    sheeps = query.all()
    today_date = datetime.now(UTC).date()
    return render_template('livestock/print.html', sheeps=sheeps, today_date=today_date)


@livestock_bp.route('/quick_weight', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def quick_weight():
    sheep = Sheep.query.filter_by(ear_tag=request.form.get('ear_tag').strip()).first()
    if sheep:
        new_weight = Decimal(request.form.get('weight') or '0')
        sheep.weight = new_weight
        db.session.add(WeightRecord(sheep_id=sheep.id, weight=new_weight, notes="ثبت سریع"))
        db.session.commit()
    return redirect(url_for('livestock.index'))

@livestock_bp.route('/bulk_action', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def bulk_action():
    sheep_ids = request.form.getlist('sheep_ids')
    action_type = request.form.get('bulk_action_type')

    if sheep_ids:
        with db.session.begin_nested():
            if action_type == 'change_status':
                new_status = request.form.get('new_status')
                # بروزرسانی وضعیت دام‌ها
                Sheep.query.filter(Sheep.id.in_(sheep_ids)).update({Sheep.status: new_status}, synchronize_session=False)

                if new_status == 'فروخته شده':
                    bulk_sale_price = normalize_amount_to_toman(request.form.get('bulk_sale_price'))
                    bulk_sale_date = request.form.get('bulk_sale_date')
                    sale_price = bulk_sale_price
                    sale_date = parse_smart_date(bulk_sale_date, datetime.now(UTC).date())

                    selected_sheeps = Sheep.query.filter(Sheep.id.in_(sheep_ids)).all()
                    transaction_count = 0

                    for sheep in selected_sheeps:
                        sheep.sale_price = sale_price
                        sheep.sale_date = sale_date
                        if sale_price > 0:
                            # جلوگیری از ثبت فاکتور تکراری برای یک پلاک
                            existing_tx = Transaction.query.filter(Transaction.description.ilike(f"%پلاک: {sheep.ear_tag}%"), Transaction.category == 'فروش دام').first()
                            if not existing_tx:
                                from app.models import BuyerCategory
                                buyer_name = 'فروش گروهی'
                                if sheep.buyer_category_id:
                                    bc = db.session.get(BuyerCategory, sheep.buyer_category_id)
                                    if bc: buyer_name = bc.name

                                new_tx = Transaction(
                                    t_type='درآمد', category='فروش دام', amount=sale_price, t_date=sale_date,
                                    is_archived=False, party_name=buyer_name,
                                    description=f"فروش سیستمی - پلاک: {sheep.ear_tag} - خریدار: {sheep.buyer_category.name if sheep.buyer_category else 'نامشخص'}"
                                )
                                db.session.add(new_tx)
                                db.session.flush()
                                AccountingEngine.record_sale(new_tx, include_vat=True)
                                transaction_count += 1

                    if transaction_count > 0:
                        flash(f'✅ {transaction_count} فاکتور فروش صادر و ثبت شد.', 'success')

            elif action_type == 'change_ration':
                Sheep.query.filter(Sheep.id.in_(sheep_ids)).update({Sheep.feed_ration_id: request.form.get('new_ration_id')}, synchronize_session=False)
            
            elif action_type == 'change_pen':
                Sheep.query.filter(Sheep.id.in_(sheep_ids)).update({Sheep.pen_id: request.form.get('new_pen_id')}, synchronize_session='fetch')

        db.session.commit()
        log_audit(f"عملیات گروهی {action_type} روی {len(sheep_ids)} دام")
        flash('عملیات گروهی با موفقیت انجام شد.', 'success')

    return redirect(url_for('livestock.index'))

@livestock_bp.route('/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_view_livestock')
def add_sheep():
    if request.method == 'POST':
        ear_tag = request.form.get('ear_tag').strip()
        
        # ---> رفع باگ: بازگشت به صفحه ثبت و نمایش هشدار پلاک تکراری <---
        if Sheep.query.filter_by(ear_tag=ear_tag).first():
            flash(f'خطا! پلاک {ear_tag} قبلاً در سیستم ثبت شده است.', 'danger')
            return redirect(url_for('livestock.add_sheep'))
            
        weight = Decimal(request.form.get('weight') or '0')
        purchase_price = normalize_amount_to_toman(request.form.get('purchase_price'))
        b_date_str = request.form.get('birth_date')
        
        qr = qrcode.make(f"پلاک: {ear_tag}")
        upload_dir = os.path.join('app', 'static', 'uploads', 'qrcodes')
        os.makedirs(upload_dir, exist_ok=True)
        qr_path = f"uploads/qrcodes/qr_{ear_tag}.png"
        qr.save(os.path.join('app', 'static', qr_path))

        new_sheep = Sheep(
            ear_tag=ear_tag, breed=request.form.get('breed'), gender=request.form.get('gender'), 
            weight=weight, purpose=request.form.get('purpose'), status=request.form.get('status'),
            feed_ration_id=request.form.get('feed_ration_id') or None, pen_id=request.form.get('pen_id') or None, 
            birth_date=parse_smart_date(b_date_str),
            qr_code_path=qr_path,             purchase_price=purchase_price
        )
        db.session.add(new_sheep)
        db.session.commit()
        if weight > 0: db.session.add(WeightRecord(sheep_id=new_sheep.id, weight=weight, notes="وزن اولیه"))
        db.session.commit()
        flash('دام جدید با موفقیت ثبت شد.', 'success')
        return redirect(url_for('livestock.index'))
    return render_template('livestock/add.html', rations=FeedRation.query.all(), pens=Pen.query.all(), breeds=BreedCategory.query.all(), purposes=PurposeCategory.query.all(), statuses=StatusCategory.query.all())

@livestock_bp.route('/profile/<int:id>')
@login_required
@permission_required('can_view_livestock')
def profile(id):
    today = datetime.now(UTC).date()
    sheep = Sheep.query.get_or_404(id)
    weight_history = WeightRecord.query.filter_by(sheep_id=id).order_by(WeightRecord.record_date.asc()).all()
    medical_history = MedicalRecord.query.filter_by(sheep_id=id).order_by(MedicalRecord.record_date.desc()).all()
    birth_history = BirthRecord.query.filter_by(mother_id=id).order_by(BirthRecord.birth_date.desc()).all()
    
    chart_labels = [w.record_date.strftime('%Y-%m-%d') for w in weight_history]
    chart_data = [w.weight for w in weight_history]
    
    adg, days_to_target = 0, None
    if len(weight_history) >= 2:
        w1, w2 = weight_history[-2], weight_history[-1]
        days = (w2.record_date - w1.record_date).days
        if days > 0: 
            adg = (w2.weight - w1.weight) / days * 1000
            if sheep.target_weight and adg > 0 and sheep.weight < sheep.target_weight:
                days_to_target = int((sheep.target_weight - sheep.weight) / (adg / 1000))
            
    next_heat_date = sheep.last_heat_date + timedelta(days=17) if sheep.gender == 'میش' and sheep.last_heat_date else None
            
    total_lambs_born = sum(b.lambs_count for b in birth_history)
    successful_lambs = sum(b.lambs_count for b in birth_history if b.status == 'موفق')
    dead_lambs = total_lambs_born - successful_lambs
    twins_count = sum(1 for b in birth_history if b.lambs_count >= 2)
    
    birth_stats = {'تک‌قلو': 0, 'دوقلو': 0, 'سه‌قلو و بیشتر': 0, 'مرده‌زایی/سقط': 0}
    for b in birth_history:
        if b.status != 'موفق': birth_stats['مرده‌زایی/سقط'] += 1
        else:
            if b.lambs_count == 1: birth_stats['تک‌قلو'] += 1
            elif b.lambs_count == 2: birth_stats['دوقلو'] += 1
            else: birth_stats['سه‌قلو و بیشتر'] += 1
            
    offsprings = Sheep.query.filter_by(mother_id=id).all()
    age_in_months = int((datetime.now(UTC).date() - sheep.birth_date).days / 30) if sheep.birth_date else 0

    mother = Sheep.query.get(sheep.mother_id) if sheep.mother_id else None
    father = Sheep.query.get(sheep.father_id) if sheep.father_id else None

    # --- محاسبات اقتصادی و هشدارهای هوشمند (رفع نقص متغیرهای قالب) ---
    birth_weight = float(get_setting('birth_weight', 3.5))
    weight_gained = float(sheep.weight or 0) - birth_weight
    days_alive = max((today - (sheep.birth_date or sheep.entry_date.date())).days, 1)
    daily_feed_cost = float(sheep.ration.daily_cost or 0) if sheep.ration else 0.0
    estimated_feed_cost = days_alive * daily_feed_cost
    
    # محاسبه بهای هر کیلو رشد (FCR Cost)
    fcr_cost = estimated_feed_cost / weight_gained if weight_gained > 0 else 0
    
    smart_alerts = []
    if adg < 0: 
        smart_alerts.append(f"🔴 هشدار بحرانی: این دام در بازه اخیر {abs(adg):.0f} گرم کاهش وزن روزانه داشته است!")
    
    # بررسی پرهیز دارویی
    for med in medical_history:
        if med.withdrawal_end_date and med.withdrawal_end_date > today:
            smart_alerts.append(f"⚠️ پرهیز دارویی: ذبح یا مصرف شیر تا {med.withdrawal_end_date} ممنوع است ({med.medicine_name}).")

    # استخراج تاریخچه تغییرات مخصوص این دام (Audit Trail)
    # اصلاح باگ target_id: جستجو بر اساس شماره پلاک در متن عملیات
    audit_history = AuditLog.query.filter(
        AuditLog.action.ilike(f"%{sheep.ear_tag}%")
    ).order_by(AuditLog.timestamp.desc()).all()

    timeline_events = []
    if sheep.birth_date: timeline_events.append({'date': sheep.birth_date, 'title': 'تولد', 'desc': 'تولد دام', 'icon': 'fa-baby', 'color': 'success'})
    elif sheep.entry_date: timeline_events.append({'date': sheep.entry_date.date(), 'title': 'ورود به گله', 'desc': 'ثبت اولیه', 'icon': 'fa-right-to-bracket', 'color': 'primary'})
    for w in weight_history: timeline_events.append({'date': w.record_date, 'title': f'وزن ({w.weight} کیلو)', 'desc': w.notes or 'ثبت روتین', 'icon': 'fa-weight-scale', 'color': 'info'})
    for m in medical_history: timeline_events.append({'date': m.record_date, 'title': f'درمان ({m.action_type})', 'desc': f"{m.medicine_name} - {m.notes or ''}", 'icon': 'fa-syringe', 'color': 'danger', 'img': m.photos})
    for b in birth_history: timeline_events.append({'date': b.birth_date, 'title': f'زایش ({b.lambs_count} بره)', 'desc': b.status, 'icon': 'fa-child', 'color': 'warning'})
    for mt in MatingRecord.query.filter_by(sheep_id=id).order_by(MatingRecord.mating_date.desc()).all():
        timeline_events.append({'date': mt.mating_date, 'title': f'جفت‌اندازی ({mt.mating_type})', 'desc': f"نتیجه: {mt.result}", 'icon': 'fa-heart', 'color': 'danger'})
    for us in UltrasoundRecord.query.filter_by(sheep_id=id).order_by(UltrasoundRecord.exam_date.desc()).all():
        timeline_events.append({'date': us.exam_date, 'title': f'سونوگرافی ({us.result})', 'desc': f"{us.fetus_count or '-'} جنین", 'icon': 'fa-stethoscope', 'color': 'info'})
    timeline_events.sort(key=lambda x: x['date'], reverse=True)

    matings = MatingRecord.query.filter_by(sheep_id=id).order_by(MatingRecord.mating_date.desc()).all()
    ultrasounds = UltrasoundRecord.query.filter_by(sheep_id=id).order_by(UltrasoundRecord.exam_date.desc()).all()
    quarantines = QuarantineRecord.query.filter_by(sheep_id=id).order_by(QuarantineRecord.start_date.desc()).all()

    from app.models import Medicine, BreedCategory, PurposeCategory, StatusCategory, BuyerCategory, SemenInventory, DrugInventory
    return render_template('livestock/profile.html', 
                           sheep=sheep, chart_labels=chart_labels, chart_data=chart_data, adg=adg,
                           medical_history=medical_history, birth_history=birth_history, timeline_events=timeline_events,
                           total_lambs=total_lambs_born, successful_lambs=successful_lambs, dead_lambs=dead_lambs, twins_count=twins_count,
                           birth_stats=birth_stats, offsprings=offsprings, fcr_cost=fcr_cost,
                           estimated_feed_cost=estimated_feed_cost, smart_alerts=smart_alerts, age_in_months=age_in_months,
                           days_to_target=days_to_target, next_heat_date=next_heat_date,
                           rams=Sheep.query.filter_by(gender='قوچ').all(), rations=FeedRation.query.all(),
                           pens=Pen.query.all(), medicines=Medicine.query.all(), breeds=BreedCategory.query.all(),
                           audit_history=audit_history,
                           purposes=PurposeCategory.query.all(), statuses=StatusCategory.query.all(),
                           buyer_categories=BuyerCategory.query.all(),
                           mother=mother, father=father, today_str=today.strftime('%Y-%m-%d'),
                           matings=matings, ultrasounds=ultrasounds, quarantines=quarantines,
                           semen_doses=SemenInventory.query.filter(SemenInventory.quantity_doses > 0).all(),
                           drugs_inventory=DrugInventory.query.filter(DrugInventory.stock_quantity > 0).order_by(DrugInventory.name).all())

@livestock_bp.route('/edit/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def edit_sheep(id):
    sheep = Sheep.query.get_or_404(id)
    old_status = sheep.status  # ذخیره وضعیت قبلی قبل از تغییر
    sheep.ear_tag = request.form.get('ear_tag', sheep.ear_tag)
    sheep.breed = request.form.get('breed', sheep.breed)
    sheep.gender = request.form.get('gender', sheep.gender)
    sheep.purpose = request.form.get('purpose', sheep.purpose)
    sheep.status = request.form.get('status', sheep.status)
    sheep.death_reason = request.form.get('death_reason') if sheep.status in ['تلف شده', 'مرده'] else None
    sheep.feed_ration_id = request.form.get('feed_ration_id') or None
    sheep.pen_id = request.form.get('pen_id') or None
    
    t_weight = request.form.get('target_weight')
    sheep.target_weight = Decimal(t_weight) if t_weight else None
    
    heat_str = request.form.get('last_heat_date')
    sheep.last_heat_date = parse_smart_date(heat_str)

    with db.session.begin_nested():
        # منطق ثبت خودکار فاکتور فروش در دفتر کل
        if sheep.status == 'فروخته شده':
            sheep.sale_price = normalize_amount_to_toman(request.form.get('sale_price'))

            s_date_str = request.form.get('sale_date')
            try:
                sheep.sale_date = parse_smart_date(s_date_str, datetime.now(UTC).date())
            except ValueError:
                sheep.sale_date = datetime.now(UTC).date()

            buyer_cat_id = request.form.get('buyer_category_id')
            sheep.buyer_category_id = int(buyer_cat_id) if (buyer_cat_id and buyer_cat_id.isdigit()) else None
            db.session.flush()
            
            if sheep.sale_price > 0:
                from app.models import BuyerCategory
                buyer_name = 'فروش نقدی'
                if sheep.buyer_category_id:
                    bc = db.session.get(BuyerCategory, sheep.buyer_category_id)
                    if bc: buyer_name = bc.name

                # بهبود شناسایی فاکتور با تطبیق دقیق شرح
                search_desc = f"فروش سیستمی - پلاک: {sheep.ear_tag}"
                existing_tx = Transaction.query.filter(
                    Transaction.description.ilike(f"{search_desc}%"),
                    Transaction.category == 'فروش دام'
                ).first()

                if not existing_tx:
                    new_tx = Transaction(
                        t_type='درآمد',
                        category='فروش دام',
                        amount=sheep.sale_price,
                        t_date=sheep.sale_date,
                        is_archived=False,
                        party_name=buyer_name,
                        description=f"فروش سیستمی - پلاک: {sheep.ear_tag} - خریدار: {sheep.buyer_category.name if sheep.buyer_category else 'نامشخص'}"
                    )
                    db.session.add(new_tx)
                    db.session.flush()
                    AccountingEngine.record_sale(new_tx, include_vat=True)
                    flash(f'✅ فروش دام (پلاک {sheep.ear_tag}) ثبت شد.', 'success')
                else:
                    existing_tx.amount = sheep.sale_price
                    existing_tx.t_date = sheep.sale_date
                    existing_tx.party_name = buyer_name
                    flash(f'ℹ️ فاکتور قبلی پلاک {sheep.ear_tag} بروز شد.', 'info')

            s_weight = request.form.get('sale_weight')
            if s_weight:
                sheep.weight = Decimal(s_weight)
                db.session.add(WeightRecord(sheep_id=sheep.id, weight=Decimal(s_weight), notes="وزن زمان فروش"))
        else:
            # سناریوی حیاتی: اگر دام قبلاً فروخته شده بود و حالا وضعیت تغییر کرد، فاکتور ابطال شود
            search_desc = f"فروش سیستمی - پلاک: {sheep.ear_tag}"
            old_tx = Transaction.query.filter(
                Transaction.description.ilike(f"{search_desc}%"),
                Transaction.category == 'فروش دام',
                Transaction.is_deleted == False
            ).first()
            if old_tx:
                old_tx.is_deleted = True
                flash(f'فاکتور فروش قبلی پلاک {sheep.ear_tag} به دلیل تغییر وضعیت دام ابطال شد.', 'info')

            sheep.sale_price = 0.0
            sheep.sale_date = None
     
    # ثبت سند حسابداری تلفات دام (کاهش دارایی زیستی)
    if sheep.status in ['تلف شده', 'مرده'] and old_status not in ['تلف شده', 'مرده', 'فروخته شده']:
        from app.accounting_engine import AccountingEngine
        from app.models import JournalEntry, JournalEntryLine, Account
        loss_value = sheep.purchase_price or (sheep.weight * 50000 if sheep.weight else 500000)
        acc_expense = Account.query.filter_by(code='5010').first()
        acc_asset = Account.query.filter_by(code='1200').first()
        if acc_expense and acc_asset:
            entry = JournalEntry(
                entry_number=AccountingEngine.generate_entry_number(),
                date=datetime.now(UTC).date(),
                description=f"تلفات دام - پلاک {sheep.ear_tag} - {sheep.death_reason or 'نامشخص'}"
            )
            db.session.add(entry)
            db.session.flush()
            db.session.add(JournalEntryLine(journal_entry_id=entry.id, account_id=acc_expense.id, debit=loss_value, credit=0.0, description=f"هزینه تلفات دام پلاک {sheep.ear_tag}"))
            db.session.add(JournalEntryLine(journal_entry_id=entry.id, account_id=acc_asset.id, debit=0.0, credit=loss_value, description=f"کاهش دارایی زیستی - پلاک {sheep.ear_tag}"))

    log_audit(f"ویرایش اطلاعات پروفایل دام پلاک {sheep.ear_tag}") 
    db.session.commit()
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/add_weight/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_weight(id):
    sheep = Sheep.query.get_or_404(id)
    new_weight = float(request.form.get('weight'))
    date_str = request.form.get('record_date')
    r_date = parse_smart_date(date_str, datetime.now(UTC).date())
    
    # دریافت نمره BCS
    bcs_val = request.form.get('bcs')
    bcs = float(bcs_val) if bcs_val else None
    
    sheep.weight = new_weight
    db.session.add(WeightRecord(sheep_id=id, weight=new_weight, bcs=bcs, notes=request.form.get('notes'), record_date=r_date))
    db.session.commit()
    
    # اعمال اتوماتیک تغییر جیره اگر دام لاغر است (BCS زیر 2.5)
    if bcs and bcs <= 2.5 and sheep.status != 'بیمار':
        flash('هشدار: نمره وضعیت بدنی (BCS) دام پایین است. سیستم پیشنهاد می‌کند جیره پرانرژی (پرواری) به این دام اختصاص دهید.', 'warning')
        
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/add_medical/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_medical(id):
    r_date = parse_smart_date(request.form.get('record_date'), datetime.now(UTC).date())
    n_date = parse_smart_date(request.form.get('next_date'))
    w_date = parse_smart_date(request.form.get('withdrawal_end_date'))
    
    action_type = request.form.get('action_type', 'درمان')
    drug_id = request.form.get('drug_id', type=int) or None
    
    record = MedicalRecord(
        sheep_id=id, action_type=action_type, 
        medicine_name=request.form.get('medicine_name', 'نامشخص'), 
        drug_id=drug_id,
        notes=request.form.get('notes'), record_date=r_date, 
        next_date=n_date, withdrawal_end_date=w_date
    )
    db.session.add(record)
    
    if drug_id:
        drug = DrugInventory.query.get(drug_id)
        if drug and drug.stock_quantity > 0:
            qty = float(request.form.get('quantity_used', 1))
            drug.stock_quantity = max(0, float(drug.stock_quantity) - qty)
    
    db.session.commit()
    
    photos = request.files.getlist('photos')
    upload_folder = os.path.join('app', 'static', 'uploads', 'medical')
    os.makedirs(upload_folder, exist_ok=True)
    for photo in photos:
        if photo and photo.filename != '':
            filename = f"{int(time.time())}_{secure_filename(photo.filename)}"
            photo.save(os.path.join(upload_folder, filename))
            db.session.add(MedicalPhoto(medical_record_id=record.id, image_path=f"uploads/medical/{filename}"))
    db.session.commit()
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/add_birth/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_birth(id):
    r_date = parse_smart_date(request.form.get('record_date'), datetime.now(UTC).date())
    father_id = request.form.get('father_id')
    lambs_count = int(request.form.get('lambs_count', 1))
    status = request.form.get('status', 'موفق')
    
    db.session.add(BirthRecord(mother_id=id, father_id=father_id or None, lambs_count=lambs_count, status=status, notes=request.form.get('notes'), birth_date=r_date))
    db.session.commit()
    
    if status == 'موفق':
        mother = Sheep.query.get(id)
        b_weight = float(get_setting('birth_weight', 3.5))
        lamb_value = Decimal(str(get_setting('market_price', '0'))) * Decimal(str(b_weight)) if get_setting('market_price', '0') != '0' else Decimal('500000')
        from app.models import JournalEntry, JournalEntryLine, Account
        entry = JournalEntry(
            entry_number=AccountingEngine.generate_entry_number(),
            date=r_date,
            description=f"زایش دام - پلاک مادر: {mother.ear_tag} - تعداد بره: {lambs_count}"
        )
        db.session.add(entry)
        db.session.flush()
        acc_asset = Account.query.filter_by(code='1200').first()
        acc_income = Account.query.filter_by(code='4010').first()
        if acc_asset and acc_income:
            for i in range(lambs_count):
                tag = f"LMB-{mother.ear_tag}-{random.randint(100, 9999)}"
                db.session.add(Sheep(ear_tag=tag, breed=mother.breed, gender="نامشخص", weight=b_weight, status="زنده و سالم", purpose="پرواربندی", birth_date=r_date, mother_id=mother.id, father_id=father_id or None))
            db.session.add(JournalEntryLine(journal_entry_id=entry.id, account_id=acc_asset.id, debit=lamb_value * lambs_count, credit=0.0, description=f"افزایش دارایی زیستی - {lambs_count} راس بره متولد شده از {mother.ear_tag}"))
            db.session.add(JournalEntryLine(journal_entry_id=entry.id, account_id=acc_income.id, debit=0.0, credit=lamb_value * lambs_count, description=f"درآمد حاصل از زایش - {lambs_count} راس بره"))
        else:
            for i in range(lambs_count):
                tag = f"LMB-{mother.ear_tag}-{random.randint(100, 9999)}"
                db.session.add(Sheep(ear_tag=tag, breed=mother.breed, gender="نامشخص", weight=b_weight, status="زنده و سالم", purpose="پرواربندی", birth_date=r_date, mother_id=mother.id, father_id=father_id or None))
        db.session.commit()
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def delete_sheep(id):
    sheep = Sheep.query.get_or_404(id)
    sheep.is_deleted = True
    sheep.status = 'حذف شده'
    db.session.commit()
    flash(f'دام پلاک {sheep.ear_tag} از لیست خارج شد.', 'warning')
    return redirect(url_for('livestock.index'))

@livestock_bp.route('/vet_queue')
@login_required
@permission_required('can_view_livestock')
def vet_queue():
    today = datetime.now(UTC).date()
    today = datetime.now(UTC).date()
    sick_sheep = Sheep.query.filter_by(status='بیمار').all()
    due_meds = MedicalRecord.query.filter(MedicalRecord.next_date != None, MedicalRecord.next_date <= today).all()
    due_ids = [m.sheep_id for m in due_meds]
    med_sheep = Sheep.query.filter(Sheep.id.in_(due_ids)).all() if due_ids else []
    newborns = [s for s in Sheep.query.filter(Sheep.birth_date >= today - timedelta(days=7)).all() if not MedicalRecord.query.filter_by(sheep_id=s.id, medicine_name="چکاپ سلامت نوزاد").first()]
    return render_template('livestock/vet_queue.html', sick_sheep=sick_sheep, due_meds=due_meds, newborns=newborns, templates=TreatmentTemplate.query.all(), today=today)

@livestock_bp.route('/apply_protocol/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def apply_protocol(id):
    template = TreatmentTemplate.query.get_or_404(request.form.get('template_id'))
    today = datetime.now(UTC).date()
    for med in template.medicines.split(','):
        db.session.add(MedicalRecord(sheep_id=id, action_type="پروتکل", medicine_name=med.strip(), record_date=today, operator="سیستم", notes=f"پروتکل: {template.name}"))
    sheep = Sheep.query.get(id)
    if sheep.status == 'بیمار': sheep.status = 'تحت درمان'
    db.session.commit()
    flash(f"پروتکل روی دام اعمال شد.", "success")
    return redirect(request.referrer)

@livestock_bp.route('/medical')
@login_required
@permission_required('can_view_livestock')
def medical_overview():
    today = datetime.now(UTC).date()
    sick_sheep = Sheep.query.filter_by(status='بیمار').all()
    upcoming_meds = MedicalRecord.query.filter(MedicalRecord.next_date != None, MedicalRecord.next_date <= today + timedelta(days=7)).order_by(MedicalRecord.next_date.asc()).all()
    return render_template('livestock/medical.html', sick_sheep=sick_sheep, upcoming_meds=upcoming_meds, today=today)

@livestock_bp.route('/genetics')
@login_required
@permission_required('can_view_livestock')
def genetics():
    return redirect(url_for('livestock.breeding'))

@livestock_bp.route('/ai-health')
@login_required
@permission_required('can_view_livestock')
def ai_health():
    """AI Health Risk Prediction - statistical anomaly detection for livestock"""
    from statistics import mean, stdev
    from datetime import timedelta

    today = datetime.now(UTC).date()
    sheep_list = Sheep.query.filter(Sheep.is_deleted == False, Sheep.status.in_(['زنده و سالم', 'بیمار', 'تحت درمان'])).all()
    predictions = []

    for sheep in sheep_list:
        weights = WeightRecord.query.filter_by(sheep_id=sheep.id).order_by(WeightRecord.record_date.asc()).all()
        meds = MedicalRecord.query.filter_by(sheep_id=sheep.id).order_by(MedicalRecord.record_date.desc()).all()

        risk_score = 0.0
        factors = []

        # 1. Weight trend anomaly (sudden drop detection)
        if len(weights) >= 3:
            recent = [float(w.weight) for w in weights[-3:]]
            older = [float(w.weight) for w in weights[:-3]]
            if older:
                older_avg = mean(older)
                recent_avg = mean(recent)
                drop_pct = ((recent_avg - older_avg) / older_avg) * 100
                if drop_pct < -5:
                    risk_score += 30
                    factors.append(f"کاهش وزن شدید ({drop_pct:.1f}%)")
                elif drop_pct < -2:
                    risk_score += 15
                    factors.append(f"کاهش وزن ({drop_pct:.1f}%)")

        # 2. BCS anomaly (if BCS data exists)
        bcs_values = [float(w.bcs) for w in weights if w.bcs]
        if len(bcs_values) >= 2:
            last_bcs = bcs_values[-1]
            if last_bcs < 2:
                risk_score += 20
                factors.append("BCS پایین (کمتر از 2)")

        # 3. Medical history recency
        recent_meds = [m for m in meds if m.record_date and (today - m.record_date).days <= 14]
        if recent_meds:
            risk_score += 25
            factors.append(f"{len(recent_meds)} مورد درمان در ۱۴ روز اخیر")

        # 4. No weight record for > 30 days
        if weights:
            last_weight_date = weights[-1].record_date
            if last_weight_date and (today - last_weight_date).days > 30:
                risk_score += 10
                factors.append("آخرین وزن‌کشی بیش از ۳۰ روز پیش")
        else:
            risk_score += 15
            factors.append("هیچ رکورد وزنی ثبت نشده")

        # 5. Age factor (very young or very old)
        if sheep.birth_date:
            age_months = (today.year - sheep.birth_date.year) * 12 + (today.month - sheep.birth_date.month)
            if age_months < 3:
                risk_score += 10
                factors.append("سن کم (زیر ۳ ماه)")
            elif age_months > 96:
                risk_score += 15
                factors.append("سن بالا (بالای ۸ سال)")

        # 6. Current status
        if sheep.status == 'بیمار':
            risk_score += 35
            factors.append("بیمار")
        elif sheep.status == 'تحت درمان':
            risk_score += 20
            factors.append("تحت درمان")

        level = "safe"
        badge_class = "bg-success"
        if risk_score >= 50:
            level = "critical"
            badge_class = "bg-danger"
        elif risk_score >= 25:
            level = "warning"
            badge_class = "bg-warning text-dark"

        predictions.append({
            'sheep': sheep,
            'risk_score': min(risk_score, 100),
            'level': level,
            'badge_class': badge_class,
            'factors': factors[:3],
            'last_weight': weights[-1].weight if weights else 0,
            'last_weight_date': weights[-1].record_date if weights else '-',
            'last_med': meds[0].medicine_name if meds else '-',
        })

    predictions.sort(key=lambda x: x['risk_score'], reverse=True)
    critical_count = sum(1 for p in predictions if p['level'] == 'critical')
    warning_count = sum(1 for p in predictions if p['level'] == 'warning')
    safe_count = sum(1 for p in predictions if p['level'] == 'safe')

    return render_template('livestock/ai_health.html',
        predictions=predictions, critical_count=critical_count,
        warning_count=warning_count, safe_count=safe_count, total=len(predictions))

@livestock_bp.route('/breeding')
@login_required
@permission_required('can_view_livestock')
def breeding():
    """صفحه یکپارچه مدیریت تولیدمثل و اصلاح نژاد (ادغام هوش ژنتیک + پیشنهاد جفت‌گیری)"""
    from sqlalchemy import func as sf
    today = datetime.now(UTC).date()
    today_str = str(today.year)

    # ---- بخش ۱: هوش ژنتیک (رتبه‌بندی دام‌های برتر) ----
    rams_all = Sheep.query.filter(Sheep.gender.in_(['قوچ', 'بره نر']), Sheep.is_deleted == False).all()
    top_rams = []
    for ram in rams_all:
        offsprings = Sheep.query.filter(Sheep.father_id == ram.id).count()
        if offsprings == 0:
            continue
        twins = BirthRecord.query.filter(BirthRecord.father_id == ram.id, BirthRecord.birth_type == 'دوقلو').count()
        top_rams.append({'ram': ram, 'offsprings': offsprings, 'twins': twins})
    top_rams.sort(key=lambda x: (x['twins'], x['offsprings']), reverse=True)
    top_rams = top_rams[:5]

    ewes_all = Sheep.query.filter(Sheep.gender.in_(['میش', 'بره ماده']), Sheep.is_deleted == False).all()
    top_ewes = []
    for ewe in ewes_all:
        successful_births = BirthRecord.query.filter(BirthRecord.mother_id == ewe.id).count()
        if successful_births == 0:
            continue
        top_ewes.append({'ewe': ewe, 'successful_births': successful_births})
    top_ewes.sort(key=lambda x: x['successful_births'], reverse=True)
    top_ewes = top_ewes[:5]

    # ---- بخش ۲: پیشنهاد جفت‌گیری (وضعیت فعلی) ----
    rams = Sheep.query.filter(Sheep.gender.in_(['قوچ', 'بره نر']), Sheep.status.in_(['زنده و سالم', 'آبستن'])).all()
    ewes = Sheep.query.filter(Sheep.gender.in_(['میش', 'بره ماده']), Sheep.status.in_(['زنده و سالم'])).all()
    total_rams = len(rams)
    total_ewes = len(ewes)
    avg_weight_rams = db.session.query(sf.avg(Sheep.weight)).filter(Sheep.gender.in_(['قوچ', 'بره نر']), Sheep.status.in_(['زنده و سالم', 'آبستن'])).scalar() or 0
    avg_weight_ewes = db.session.query(sf.avg(Sheep.weight)).filter(Sheep.gender.in_(['میش', 'بره ماده']), Sheep.status.in_(['زنده و سالم'])).scalar() or 0

    return render_template('livestock/breeding.html',
        top_rams=top_rams, top_ewes=top_ewes, today_str=today_str,
        rams=rams, ewes=ewes, total_rams=total_rams, total_ewes=total_ewes,
        avg_weight_rams=round(float(avg_weight_rams), 1),
        avg_weight_ewes=round(float(avg_weight_ewes), 1), today=today)

@livestock_bp.route('/mating-suggestion')
@login_required
@permission_required('can_view_livestock')
def mating_suggestion():
    return redirect(url_for('livestock.breeding'))

@livestock_bp.route('/mark_healthy/<int:id>')
@login_required
@permission_required('can_view_livestock')
def mark_healthy(id):
    sheep = Sheep.query.get_or_404(id)
    sheep.status = 'زنده و سالم'
    db.session.commit()
    flash(f"دام {sheep.ear_tag} ترخیص شد.", "success")
    return redirect(request.referrer)

@livestock_bp.route('/mark_med_done/<int:med_id>')
@login_required
@permission_required('can_view_livestock')
def mark_med_done(med_id):
    old_med = MedicalRecord.query.get_or_404(med_id)
    db.session.add(MedicalRecord(sheep_id=old_med.sheep_id, action_type=old_med.action_type, medicine_name=old_med.medicine_name, record_date=datetime.now(UTC).date(), operator="سیستم", notes="تکرار نوبت قبلی انجام شد."))
    old_med.next_date = None
    db.session.commit()
    flash(f"داروی {old_med.medicine_name} ثبت شد و از صف حذف گردید.", "success")
    return redirect(request.referrer)

@livestock_bp.route('/mark_newborn_checked/<int:id>')
@login_required
@permission_required('can_view_livestock')
def mark_newborn_checked(id):
    db.session.add(MedicalRecord(sheep_id=id, action_type="ویزیت", medicine_name="چکاپ سلامت نوزاد", record_date=datetime.now(UTC).date(), operator="سیستم", notes="ویزیت اولیه ثبت شد."))
    db.session.commit()
    flash("چکاپ نوزاد ثبت شد و از صف حذف گردید.", "success")
    return redirect(request.referrer)

@livestock_bp.route('/toggle_star/<int:id>', methods=['POST'])
@rate_limit(limit=30, per=60)
@login_required
@permission_required('can_view_livestock')
def toggle_star(id):
    sheep = Sheep.query.get_or_404(id)
    sheep.is_starred = not sheep.is_starred
    db.session.commit()
    return jsonify({'success': True, 'is_starred': sheep.is_starred})

@livestock_bp.route('/passport/<ear_tag>')
def public_passport(ear_tag):
    sheep = Sheep.query.filter_by(ear_tag=ear_tag).first_or_404()
    from app.models import SystemSetting
    # فقط تنظیمات غیرحساس برای نمایش عمومی
    safe_keys = ['farm_name', 'farm_logo_path', 'currency_unit']
    safe_settings = {}
    for key in safe_keys:
        s = SystemSetting.query.filter_by(key=key).first()
        if s: safe_settings[key] = s.value
    
    # محاسبه سن
    age = "نامشخص"
    if sheep.birth_date:
        age_days = (datetime.now(UTC).date() - sheep.birth_date).days
        age = f"{age_days // 30} ماه"
        
    return render_template('livestock/passport.html', sheep=sheep, age=age, settings=safe_settings)


# ==========================================
# مانیتورینگ هوشمند و اینترنت اشیا (IoT) بهاربندها
# ==========================================
@livestock_bp.route('/pens')
@login_required
@permission_required('can_view_livestock')
def pen_dashboard():
    from app.models import Pen, SensorData
    pens = Pen.query.all()
    selected_pen_id = request.args.get('pen_id', type=int)
    
    selected_pen = None
    stats = {'total': 0, 'avg_weight': 0.0, 'capacity': 0, 'fill_percentage': 0.0}
    sensor = None
    thi = 0
    thi_status = "نرمال"
    sheeps_pagination = None
    
    # متغیرهای فیلتر
    search_q = request.args.get('search', '').strip()
    gender_q = request.args.get('gender', 'همه')
    breed_q = request.args.get('breed', 'همه')
    status_q = request.args.get('status', 'همه')
    starred_q = request.args.get('starred')
    min_w = request.args.get('min_weight', type=float)
    max_w = request.args.get('max_weight', type=float)
    
    if not selected_pen_id and pens:
        selected_pen_id = pens[0].id
        
    if selected_pen_id:
        selected_pen = Pen.query.get_or_404(selected_pen_id)
        
        # --- محاسبات آماری نمودارها (سریع) ---
        active_sheeps = [s for s in selected_pen.sheep_list if s.status not in ['تلف شده', 'مرده', 'فروخته شده']]
        stats['total'] = len(active_sheeps)
        total_weight = sum(s.weight for s in active_sheeps if s.weight)
        stats['avg_weight'] = (total_weight / stats['total']) if stats['total'] > 0 else 0
        stats['capacity'] = selected_pen.capacity
        stats['fill_percentage'] = (stats['total'] / stats['capacity'] * 100) if stats['capacity'] > 0 else 0
        
        stats['genders'], stats['breeds'], stats['statuses'] = {}, {}, {}
        for s in active_sheeps:
            g = 'بره' if 'بره' in s.gender else s.gender
            stats['genders'][g] = stats['genders'].get(g, 0) + 1
            breed = s.breed or 'نامشخص'
            stats['breeds'][breed] = stats['breeds'].get(breed, 0) + 1
            stats['statuses'][s.status] = stats['statuses'].get(s.status, 0) + 1
            
        # دیتای سنسور (IoT)
        sensor = SensorData.query.filter_by(pen_id=selected_pen.id).order_by(SensorData.recorded_at.desc()).first()

        if sensor:
            # شاخص THI (تبدیل به float چون Numeric دیتابیس Decimal برمی‌گرداند)
            t = float(sensor.temperature)
            rh = float(sensor.humidity)
            thi = (0.8 * t) + ((rh / 100) * (t - 14.4)) + 46.4
            if thi < 72: thi_status = "نرمال و راحت"
            elif thi < 78: thi_status = "تنش حرارتی خفیف"
            elif thi < 82: thi_status = "تنش شدید (خطر)"
            else: thi_status = "بحران گرمایی!"
        else:
            thi = 0
            thi_status = "بدون داده"

        # --- رفع کندی: سیستم فیلتر بک اند و صفحه بندی ---
        query = Sheep.query.filter_by(pen_id=selected_pen.id)
        if search_q: query = query.filter(Sheep.ear_tag.ilike(f"%{search_q}%"))
        if gender_q != 'همه': query = query.filter(Sheep.gender == gender_q)
        if breed_q != 'همه': query = query.filter(Sheep.breed == breed_q)
        if status_q != 'همه': query = query.filter(Sheep.status == status_q)
        if starred_q == '1': query = query.filter(Sheep.is_starred.is_(True))
        if min_w is not None: query = query.filter(Sheep.weight >= min_w)
        if max_w is not None: query = query.filter(Sheep.weight <= max_w)
        
        page = request.args.get('page', 1, type=int)
        sheeps_pagination = query.order_by(Sheep.id.desc()).paginate(page=page, per_page=50)
            
    return render_template('livestock/pens.html', 
                           pens=pens, selected_pen=selected_pen, 
                           stats=stats, sensor=sensor, thi=thi, thi_status=thi_status, 
                           sheeps=sheeps_pagination,
                           current_search=search_q, current_gender=gender_q, 
                           current_breed=breed_q, current_status=status_q, current_starred=starred_q,
                           current_min_w=min_w, current_max_w=max_w, breeds=BreedCategory.query.all(),
                           rations=FeedRation.query.all(), statuses=StatusCategory.query.all())

@livestock_bp.route('/api/pen/<int:pen_id>/sensor')
@login_required
@permission_required('can_view_livestock')
def get_pen_sensor_api(pen_id):
    """API مخصوص به‌روزرسانی خودکار کارت‌های سنسور در پس‌زمینه"""
    from app.models import SensorData
    sensor = SensorData.query.filter_by(pen_id=pen_id).order_by(SensorData.recorded_at.desc()).first()
    if not sensor:
        return jsonify({"status": "no_data"}), 200
    
    t = float(sensor.temperature)
    rh = float(sensor.humidity)
    thi = (0.8 * t) + ((rh / 100) * (t - 14.4)) + 46.4
    
    if thi < 72: thi_status = "نرمال و راحت"
    elif thi < 78: thi_status = "تنش حرارتی خفیف"
    elif thi < 82: thi_status = "تنش شدید (خطر)"
    else: thi_status = "بحران گرمایی!"
    
    return jsonify({
        "status": "success",
        "temperature": t,
        "humidity": rh,
        "thi": round(thi, 1),
        "thi_status": thi_status
    })

@livestock_bp.route('/maintenance/cleanup_weights', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def cleanup_weights():
    """حذف رکوردهای وزن‌کشی قدیمی‌تر از ۲ سال جهت بهینه‌سازی دیتابیس"""
    if current_user.role != 'مدیر':
        flash('فقط مدیر کل دسترسی به عملیات نگهداری و سبک‌سازی دیتابیس دارد.', 'danger')
        return redirect(url_for('dashboard.index'))

    cutoff_date = datetime.now(UTC).date() - timedelta(days=730)
    deleted_count = WeightRecord.query.filter(WeightRecord.record_date < cutoff_date).delete()
    db.session.commit()

    flash(f'عملیات سبک‌سازی با موفقیت انجام شد. تعداد {deleted_count} رکورد وزن‌کشی قدیمی (بیش از ۲ سال) از سیستم حذف گردید.', 'success')
    return redirect(url_for('dashboard.settings'))

@livestock_bp.route('/api/sheep_search')
@login_required
@permission_required('can_view_livestock')
def sheep_search():
    q = request.args.get('q', '').strip()
    gender_filter = request.args.get('gender', '').strip()
    if not q or len(q) < 1:
        return jsonify([])
    results = Sheep.query.filter(
        Sheep.is_deleted == False,
        Sheep.ear_tag.ilike(f'%{q}%')
    )
    if gender_filter == 'ماده':
        results = results.filter(Sheep.gender.in_(['میش', 'بره ماده']))
    elif gender_filter == 'نر':
        results = results.filter(Sheep.gender.in_(['قوچ', 'بره نر']))
    results = results.order_by(Sheep.ear_tag).limit(20).all()
    return jsonify([{
        'id': s.id, 'ear_tag': s.ear_tag, 'breed': s.breed or 'نامشخص',
        'gender': s.gender, 'pen': s.pen.name if s.pen else 'بدون بهاربند',
        'status': s.status, 'weight': float(s.weight) if s.weight else 0
    } for s in results])

@livestock_bp.route('/api/mating_suggestions/<int:sheep_id>')
@login_required
@permission_required('can_view_livestock')
def mating_suggestions(sheep_id):
    ewe = Sheep.query.get_or_404(sheep_id)
    if ewe.gender not in ('میش', 'بره ماده'):
        return jsonify({'error': 'not a ewe'}), 400
    suggestions = []
    rams = Sheep.query.filter(Sheep.is_deleted == False, Sheep.gender == 'قوچ').all()
    for ram in rams:
        total = MatingRecord.query.filter_by(male_id=ram.id).count()
        successful = MatingRecord.query.filter_by(male_id=ram.id, result='آبستن').count()
        failed = MatingRecord.query.filter_by(male_id=ram.id, result='خالی').count()
        rate = (successful / (successful + failed) * 100) if (successful + failed) > 0 else None
        # همخونی
        inbred = (ewe.father_id and ram.father_id and ewe.father_id == ram.father_id) or \
                 (ewe.mother_id and ram.mother_id and ewe.mother_id == ram.mother_id)
        suggestions.append({
            'id': ram.id, 'name': ram.ear_tag, 'breed': ram.breed or 'نامشخص',
            'total': total, 'successful': successful, 'failed': failed,
            'success_rate': round(rate, 1) if rate is not None else None,
            'inbred': inbred, 'status': ram.status
        })
    suggestions.sort(key=lambda x: (x['inbred'], -(x['success_rate'] or 0)))
    # تاریخ پیشنهادی
    recommended_date = None
    if ewe.last_heat_date:
        recommended_date = (ewe.last_heat_date + timedelta(days=17)).strftime('%Y-%m-%d')
    return jsonify({'suggestions': suggestions[:5], 'recommended_date': recommended_date})

# ============================================================
# تولیدمثل: جفت‌اندازی، سونوگرافی، تقویم آبستنی
# ============================================================

@livestock_bp.route('/add_mating/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_mating(id):
    sheep = Sheep.query.get_or_404(id)
    male_id = request.form.get('male_id') or None
    semen_id = request.form.get('semen_id', type=int) or None
    mating = MatingRecord(
        sheep_id=id,
        male_id=int(male_id) if male_id else None,
        mating_date=parse_smart_date(request.form.get('mating_date')),
        mating_type=request.form.get('mating_type', 'طبیعی'),
        semen_id=semen_id,
        notes=request.form.get('notes')
    )
    db.session.add(mating)
    if semen_id:
        semen = SemenInventory.query.get(semen_id)
        if semen and semen.quantity_doses > 0:
            semen.quantity_doses -= 1
    sheep.last_heat_date = mating.mating_date
    if sheep.status == 'زنده و سالم':
        sheep.status = 'جفت‌اندازی شده'
    db.session.commit()
    log_audit(f"ثبت جفت‌اندازی {sheep.ear_tag} ({mating.mating_type})")
    flash('جفت‌اندازی ثبت شد.', 'success')
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/add_ultrasound/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_ultrasound(id):
    sheep = Sheep.query.get_or_404(id)
    result = request.form.get('result')
    exam_date = parse_smart_date(request.form.get('exam_date'))
    fetus_count = request.form.get('fetus_count', type=int) or None
    gestational_days = request.form.get('gestational_days', type=int) or None
    due_date = None
    if result == 'آبستن' and gestational_days:
        due_date = exam_date + timedelta(days=(147 - gestational_days))
    elif result == 'آبستن' and request.form.get('due_date'):
        due_date = parse_smart_date(request.form.get('due_date'))
    us = UltrasoundRecord(
        sheep_id=id, exam_date=exam_date,
        exam_type=request.form.get('exam_type', 'روتین'),
        result=result, fetus_count=fetus_count,
        gestational_days=gestational_days,
        due_date=due_date, notes=request.form.get('notes')
    )
    db.session.add(us)
    db.session.flush()
    # آپلود تصاویر سونوگرافی
    images = request.files.getlist('us_images')
    for img in images:
        if img and img.filename:
            filename = f"us_{id}_{int(time.time())}_{secure_filename(img.filename)}"
            os.makedirs(os.path.join('app', 'static', 'uploads', 'medical'), exist_ok=True)
            img.save(os.path.join('app', 'static', 'uploads', 'medical', filename))
            db.session.add(UltrasoundImage(ultrasound_id=us.id, image_path=f"uploads/medical/{filename}"))
    if result == 'آبستن':
        sheep.status = 'آبستن'
        # به‌روزرسانی آخرین جفت‌اندازی مرتبط
        last_mating = MatingRecord.query.filter_by(sheep_id=id, result='منتظر نتیجه').order_by(MatingRecord.id.desc()).first()
        if last_mating:
            last_mating.result = 'آبستن'
            last_mating.result_date = exam_date
            last_mating.confirmed_by = 'سونوگرافی'
    elif result == 'خالی':
        last_mating = MatingRecord.query.filter_by(sheep_id=id, result='منتظر نتیجه').order_by(MatingRecord.id.desc()).first()
        if last_mating:
            last_mating.result = 'خالی'
            last_mating.result_date = exam_date
            last_mating.confirmed_by = 'سونوگرافی'
    db.session.commit()
    log_audit(f"ثبت سونوگرافی {sheep.ear_tag}: {result}")
    flash(f'نتیجه سونوگرافی ثبت شد: {result}', 'success')
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/breeding_calendar')
@login_required
@permission_required('can_view_livestock')
def breeding_calendar():
    import jdatetime
    now_j = jdatetime.datetime.now()
    year = request.args.get('year', now_j.year, type=int)
    month = request.args.get('month', now_j.month, type=int)
    view = request.args.get('view', 'month')
    pen_filter = request.args.get('pen_id', type=int) or None
    event_filter = request.args.get('event_type', '')
    is_print = request.args.get('print', type=int) == 1

    month_names = {1:'فروردین',2:'اردیبهشت',3:'خرداد',4:'تیر',5:'مرداد',6:'شهریور',7:'مهر',8:'آبان',9:'آذر',10:'دی',11:'بهمن',12:'اسفند'}
    j_first = jdatetime.date(year, month, 1)
    j_last = jdatetime.date(year + 1, 1, 1) if month == 12 else jdatetime.date(year, month + 1, 1)
    g_first = j_first.togregorian()
    g_last = j_last.togregorian()
    days_in_month = (g_last - g_first).days
    g_today = datetime.now(UTC).date()
    today_day = (g_today - g_first).days + 1 if g_first <= g_today < g_last else 0

    # فیلتر بر اساس بهاربند
    def pen_mating_q(base):
        if pen_filter:
            return base.join(Sheep, MatingRecord.sheep_id == Sheep.id).filter(Sheep.pen_id == pen_filter)
        return base
    def pen_us_q(base):
        if pen_filter:
            return base.join(Sheep, UltrasoundRecord.sheep_id == Sheep.id).filter(Sheep.pen_id == pen_filter)
        return base
    def pen_birth_q(base):
        if pen_filter:
            return base.join(Sheep, BirthRecord.mother_id == Sheep.id).filter(Sheep.pen_id == pen_filter)
        return base

    # جفت‌اندازی‌های این ماه
    matings_q = MatingRecord.query.filter(
        MatingRecord.mating_date >= g_first, MatingRecord.mating_date < g_last
    ).options(db.joinedload(MatingRecord.sheep), db.joinedload(MatingRecord.male))
    matings = pen_mating_q(matings_q).all()

    # سونوگرافی‌های این ماه
    us_q = UltrasoundRecord.query.filter(
        UltrasoundRecord.exam_date >= g_first, UltrasoundRecord.exam_date < g_last
    ).options(db.joinedload(UltrasoundRecord.sheep))
    ultrasounds = pen_us_q(us_q).all()

    # زایش‌های این ماه
    births_q = BirthRecord.query.filter(
        BirthRecord.birth_date >= g_first, BirthRecord.birth_date < g_last
    ).options(db.joinedload(BirthRecord.mother))
    births = pen_birth_q(births_q).all()

    # زایش‌های پیش‌بینی شده
    pred_q = UltrasoundRecord.query.filter(
        UltrasoundRecord.due_date >= g_first, UltrasoundRecord.due_date < g_last,
        UltrasoundRecord.result == 'آبستن'
    ).options(db.joinedload(UltrasoundRecord.sheep))
    predicted_births = pen_us_q(pred_q).all()

    # فیلتر نوع رویداد
    if event_filter == 'mating':
        ultrasounds = []; births = []; predicted_births = []
    elif event_filter == 'ultrasound':
        matings = []; births = []; predicted_births = []
    elif event_filter == 'birth':
        matings = []; ultrasounds = []; predicted_births = []
    elif event_filter == 'predicted':
        matings = []; ultrasounds = []; births = []

    # داده‌های ویو هفتگی
    weeks = []
    if view == 'week':
        current_week_start = request.args.get('week_start', type=int) or 1
        week_end = min(current_week_start + 6, days_in_month)
        weeks = [(d, current_week_start + d) for d in range(7) if current_week_start + d <= days_in_month]

    # داده‌های ویو لیستی
    all_events = []
    if view == 'list':
        for m in matings:
            all_events.append({'date': m.mating_date, 'type': 'جفت‌اندازی', 'sheep_tag': m.sheep.ear_tag if m.sheep else '?', 'detail': m.mating_type, 'color': 'danger', 'sheep_id': m.sheep_id})
        for u in ultrasounds:
            all_events.append({'date': u.exam_date, 'type': 'سونوگرافی', 'sheep_tag': u.sheep.ear_tag if u.sheep else '?', 'detail': u.result, 'color': 'info', 'sheep_id': u.sheep_id})
        for b in births:
            all_events.append({'date': b.birth_date, 'type': 'زایش', 'sheep_tag': b.mother.ear_tag if b.mother else '?', 'detail': f'{b.lambs_count} بره', 'color': 'success', 'sheep_id': b.mother_id})
        for p in predicted_births:
            all_events.append({'date': p.due_date, 'type': 'زایش پیش‌بینی', 'sheep_tag': p.sheep.ear_tag if p.sheep else '?', 'detail': 'در انتظار', 'color': 'warning', 'sheep_id': p.sheep_id})
        all_events.sort(key=lambda x: x['date'])

    # آمار قوچ‌ها
    ram_stats = []
    rams_list = Sheep.query.filter(Sheep.is_deleted == False, Sheep.gender == 'قوچ').order_by(Sheep.ear_tag).all()
    for ram in rams_list:
        total_matings = MatingRecord.query.filter_by(male_id=ram.id).count()
        if total_matings == 0:
            continue
        successful = MatingRecord.query.filter_by(male_id=ram.id, result='آبستن').count()
        failed = MatingRecord.query.filter_by(male_id=ram.id, result='خالی').count()
        pending = total_matings - successful - failed
        success_rate = (successful / (successful + failed) * 100) if (successful + failed) > 0 else None
        ram_stats.append({
            'name': ram.ear_tag, 'id': ram.id, 'total': total_matings,
            'successful': successful, 'failed': failed, 'pending': pending,
            'success_rate': success_rate
        })
    ram_stats.sort(key=lambda x: x['total'], reverse=True)

    # ===== ۱. پیش‌بینی زایش ۶ ماه آینده =====
    forecast_months = []
    forecast_counts = []
    forecast_lambs = []
    for i in range(6):
        m = now_j.month + i
        y = now_j.year
        while m > 12: m -= 12; y += 1
        fm_start = jdatetime.date(y, m, 1).togregorian()
        fm_end = (jdatetime.date(y + 1, 1, 1) if m == 12 else jdatetime.date(y, m + 1, 1)).togregorian()
        cnt = UltrasoundRecord.query.filter(
            UltrasoundRecord.due_date >= fm_start, UltrasoundRecord.due_date < fm_end,
            UltrasoundRecord.result == 'آبستن'
        ).count()
        total_lambs = db.session.query(func.coalesce(func.sum(UltrasoundRecord.fetus_count), 0)).filter(
            UltrasoundRecord.due_date >= fm_start, UltrasoundRecord.due_date < fm_end,
            UltrasoundRecord.result == 'آبستن'
        ).scalar()
        forecast_counts.append(cnt)
        forecast_lambs.append(int(total_lambs))
        forecast_months.append(month_names.get(m, ''))

    # آماده‌سازی داده‌های پایه (قبل از smart alerts)
    ready_ewes = Sheep.query.filter(
        Sheep.is_deleted == False,
        Sheep.gender.in_(['میش', 'بره ماده']),
        Sheep.status.in_(['زنده و سالم', 'آبستن']),
        Sheep.last_heat_date != None
    ).order_by(Sheep.last_heat_date.desc()).all()
    next_heats = [{'ear_tag': e.ear_tag, 'next': e.last_heat_date + timedelta(days=17), 'sheep': e} for e in ready_ewes[:10]]

    # آمار ماه (قبل از smart alerts)
    total_events = len(matings) + len(ultrasounds) + len(births) + len(predicted_births)
    pregnant_ultrasounds = [u for u in ultrasounds if u.result == 'آبستن']
    empty_ultrasounds = [u for u in ultrasounds if u.result == 'خالی']
    preg_rate = (len(pregnant_ultrasounds) / (len(pregnant_ultrasounds) + len(empty_ultrasounds)) * 100) if (len(pregnant_ultrasounds) + len(empty_ultrasounds)) > 0 else None

    # ===== ۲. Smart Alerts =====
    smart_alerts = []
    # ۲-۱. فحلی در ۳ روز آینده
    coming_heats = [e for e in ready_ewes if e.last_heat_date and 0 <= (e.last_heat_date + timedelta(days=17) - g_today).days <= 3]
    if coming_heats:
        tags = '، '.join([e.ear_tag for e in coming_heats[:5]])
        if len(coming_heats) > 5: tags += f' و {len(coming_heats)-5} رأس دیگر'
        smart_alerts.append({'type': 'heat', 'text': f'🔥 {len(coming_heats)} رأس در ۳ روز آینده فحلی دارند — {tags}', 'severity': 'warning'})
    # ۲-۲. قوچ‌های کم‌کار
    for ram in rams_list:
        two_month_ago = g_today - timedelta(days=60)
        recent = MatingRecord.query.filter_by(male_id=ram.id).filter(MatingRecord.mating_date >= two_month_ago).count()
        if recent is not None and recent <= 2:
            smart_alerts.append({'type': 'ram', 'text': f'🐏 قوچ {ram.ear_tag} فقط {recent} جفت‌اندازی در ۲ ماه داشته — بررسی سلامت', 'severity': 'danger'})
            break
    # ۲-۳. میش‌های با شکست مکرر
    for ewe in ready_ewes:
        recent_matings = MatingRecord.query.filter_by(sheep_id=ewe.id).order_by(MatingRecord.id.desc()).limit(3).all()
        failures = [m for m in recent_matings if m.result == 'خالی']
        if len(failures) >= 2:
            smart_alerts.append({'type': 'ewe', 'text': f'🐑 میش {ewe.ear_tag} از {len(recent_matings)} جفت‌اندازی پشت سر هم خالی بوده — بررسی یا حذف از گله', 'severity': 'danger'})
            break
    # ۲-۴. افت نرخ آبستنی
    if preg_rate is not None:
        prev_m = month - 1 if month > 1 else 12
        prev_y = year if month > 1 else year - 1
        prev_j_first = jdatetime.date(prev_y, prev_m, 1)
        prev_g_first = prev_j_first.togregorian()
        prev_preg = UltrasoundRecord.query.filter(
            UltrasoundRecord.exam_date >= prev_g_first, UltrasoundRecord.exam_date < g_first,
            UltrasoundRecord.result == 'آبستن'
        ).count()
        prev_empty = UltrasoundRecord.query.filter(
            UltrasoundRecord.exam_date >= prev_g_first, UltrasoundRecord.exam_date < g_first,
            UltrasoundRecord.result == 'خالی'
        ).count()
        prev_rate = (prev_preg / (prev_preg + prev_empty) * 100) if (prev_preg + prev_empty) > 0 else None
        if prev_rate and preg_rate < prev_rate - 10:
            smart_alerts.append({'type': 'rate', 'text': f'📉 نرخ آبستنی {preg_rate:.0f}٪ — نسبت به ماه قبل {prev_rate:.0f}٪ افت داشته', 'severity': 'danger'})
    # ۲-۵. دام‌های عقب افتاده از سونوگرافی
    overdue_tags = []
    for mating in MatingRecord.query.filter(MatingRecord.result == 'منتظر نتیجه').options(db.joinedload(MatingRecord.sheep)).all():
        if mating.sheep and (g_today - mating.mating_date).days >= 35:
            overdue_tags.append(mating.sheep.ear_tag)
    if overdue_tags:
        tags = '، '.join(overdue_tags[:5])
        if len(overdue_tags) > 5: tags += f' و {len(overdue_tags)-5} رأس دیگر'
        smart_alerts.append({'type': 'us', 'text': f'🔬 {len(overdue_tags)} رأس بیش از ۳۵ روز از جفت‌اندازی گذشته — نیاز به سونوگرافی: {tags}', 'severity': 'warning'})

    # ===== ۳. نقشه حرارتی باروری (۱۲ ماه) =====
    fertility_months = []
    fertility_rates = []
    fertility_counts = []
    for i in range(11, -1, -1):
        m = now_j.month - i
        y = now_j.year
        while m < 1: m += 12; y -= 1
        while m > 12: m -= 12; y += 1
        fm_start = jdatetime.date(y, m, 1).togregorian()
        fm_end = (jdatetime.date(y + 1, 1, 1) if m == 12 else jdatetime.date(y, m + 1, 1)).togregorian()
        preg_m = UltrasoundRecord.query.filter(
            UltrasoundRecord.exam_date >= fm_start, UltrasoundRecord.exam_date < fm_end,
            UltrasoundRecord.result == 'آبستن'
        ).count()
        empty_m = UltrasoundRecord.query.filter(
            UltrasoundRecord.exam_date >= fm_start, UltrasoundRecord.exam_date < fm_end,
            UltrasoundRecord.result == 'خالی'
        ).count()
        rate = (preg_m / (preg_m + empty_m) * 100) if (preg_m + empty_m) > 0 else None
        fertility_months.append(month_names.get(m, ''))
        fertility_rates.append(rate)
        fertility_counts.append(preg_m + empty_m)

    # ===== ۴. گانت چارت (میش‌های آبستن) =====
    gantt_data = []
    pregnant_q = Sheep.query.filter(Sheep.is_deleted == False, Sheep.gender.in_(['میش', 'بره ماده']), Sheep.status == 'آبستن')
    if pen_filter:
        pregnant_q = pregnant_q.filter(Sheep.pen_id == pen_filter)
    for ewe in pregnant_q.order_by(Sheep.ear_tag).limit(30).all():
        last_us = UltrasoundRecord.query.filter_by(sheep_id=ewe.id, result='آبستن').order_by(UltrasoundRecord.id.desc()).first()
        due_date = last_us.due_date if last_us and last_us.due_date else None
        days_left = (due_date - g_today).days if due_date else None
        gantt_data.append({
            'ear_tag': ewe.ear_tag, 'id': ewe.id,
            'confirm_date': last_us.exam_date if last_us else None,
            'due_date': due_date, 'days_left': days_left,
            'fetus_count': last_us.fetus_count if last_us else None
        })
    gantt_data.sort(key=lambda x: x['days_left'] if x['days_left'] is not None else 999)

    # میش‌های آماده جفت‌اندازی
    # داده‌های مورد نیاز
    rams = rams_list
    semen_doses = SemenInventory.query.filter(SemenInventory.quantity_doses > 0).order_by(SemenInventory.ram_name).all()
    pens = Pen.query.order_by(Pen.name).all()

    return render_template('livestock/breeding_calendar.html',
        year=year, month=month, view=view, pen_filter=pen_filter, event_filter=event_filter, is_print=is_print,
        month_name=month_names.get(month, ''), month_names=month_names, days_in_month=days_in_month,
        g_first=g_first, g_today=g_today, today_day=today_day,
        matings=matings, ultrasounds=ultrasounds, births=births, predicted_births=predicted_births,
        next_heats=next_heats, rams=rams, semen_doses=semen_doses, pens=pens,
        total_events=total_events, preg_rate=preg_rate,
        preg_count=len(pregnant_ultrasounds), empty_count=len(empty_ultrasounds), birth_count=len(births),
        weeks=weeks, all_events=all_events if view == 'list' else [], ram_stats=ram_stats,
        forecast_months=forecast_months, forecast_counts=forecast_counts, forecast_lambs=forecast_lambs,
        smart_alerts=smart_alerts,
        fertility_months=fertility_months, fertility_rates=fertility_rates, fertility_counts=fertility_counts,
        gantt_data=gantt_data)

@livestock_bp.route('/bulk_mating', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def bulk_mating():
    sheep_ids = request.form.getlist('sheep_ids')
    male_id = request.form.get('male_id')
    mating_date = parse_smart_date(request.form.get('mating_date'))
    mating_type = request.form.get('mating_type', 'طبیعی')
    if not sheep_ids or not mating_date:
        flash('لطفاً دام‌ها و تاریخ را انتخاب کنید.', 'danger')
        return redirect(url_for('livestock.breeding_calendar'))
    count = 0
    for sid in sheep_ids:
        sid = sid.strip()
        if not sid.isdigit(): continue
        sheep = Sheep.query.get(int(sid))
        if not sheep or sheep.is_deleted: continue
        db.session.add(MatingRecord(
            sheep_id=int(sid),
            male_id=int(male_id) if male_id else None,
            mating_date=mating_date,
            mating_type=mating_type,
            notes=request.form.get('notes')
        ))
        sheep.last_heat_date = mating_date
        if sheep.status == 'زنده و سالم':
            sheep.status = 'جفت‌اندازی شده'
        count += 1
    db.session.commit()
    log_audit(f"جفت‌اندازی گروهی {count} رأس ({mating_type})")
    flash(f'جفت‌اندازی گروهی برای {count} رأس ثبت شد.', 'success')
    return redirect(url_for('livestock.breeding_calendar'))

@livestock_bp.route('/ram_stats')
@login_required
@permission_required('can_view_livestock')
def ram_stats_list():
    import math
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    min_rate = request.args.get('min_rate', type=float)
    breed_filter = request.args.get('breed', '').strip()
    sort_by = request.args.get('sort', 'total').strip()

    q = Sheep.query.filter(Sheep.is_deleted == False, Sheep.gender == 'قوچ')
    if search:
        q = q.filter(Sheep.ear_tag.ilike(f'%{search}%'))
    if breed_filter:
        q = q.filter(Sheep.breed == breed_filter)
    rams = q.order_by(Sheep.ear_tag).all()

    ram_list = []
    for ram in rams:
        total_matings = MatingRecord.query.filter_by(male_id=ram.id).count()
        successful = MatingRecord.query.filter_by(male_id=ram.id, result='آبستن').count()
        failed = MatingRecord.query.filter_by(male_id=ram.id, result='خالی').count()
        pending = total_matings - successful - failed
        success_rate = (successful / (successful + failed) * 100) if (successful + failed) > 0 else None
        ram_list.append({
            'id': ram.id, 'name': ram.ear_tag, 'breed': ram.breed or 'نامشخص',
            'total': total_matings, 'successful': successful, 'failed': failed,
            'pending': pending, 'success_rate': success_rate, 'status': ram.status
        })

    if min_rate is not None:
        ram_list = [r for r in ram_list if r['success_rate'] is not None and r['success_rate'] >= min_rate]

    if sort_by == 'rate':
        ram_list.sort(key=lambda x: (x['success_rate'] or 0), reverse=True)
    elif sort_by == 'name':
        ram_list.sort(key=lambda x: x['name'])
    else:
        ram_list.sort(key=lambda x: x['total'], reverse=True)

    total = len(ram_list)
    per_page = 50
    total_pages = max(1, math.ceil(total / per_page))
    offset = (page - 1) * per_page
    page_rams = ram_list[offset:offset + per_page]

    breeds = db.session.query(Sheep.breed).filter(
        Sheep.is_deleted == False, Sheep.gender == 'قوچ',
        Sheep.breed != None, Sheep.breed != ''
    ).distinct().order_by(Sheep.breed).all()
    breeds = [b[0] for b in breeds]

    return render_template('livestock/ram_stats.html',
        rams=page_rams, page=page, total_pages=total_pages, total=total,
        per_page=per_page, search=search, min_rate=min_rate,
        breed_filter=breed_filter, breeds=breeds, sort_by=sort_by)

@livestock_bp.route('/semen_inventory')
@login_required
@permission_required('can_view_livestock')
def semen_inventory():
    q = request.args.get('q', '').strip()
    doses = SemenInventory.query
    if q:
        doses = doses.filter(SemenInventory.ram_name.ilike(f'%{q}%') | SemenInventory.breed.ilike(f'%{q}%'))
    doses = doses.order_by(SemenInventory.expiry_date).all()
    total_doses = sum(d.quantity_doses for d in doses)
    unique_rams = len(set(d.ram_name for d in doses))
    return render_template('livestock/semen_inventory.html', doses=doses, q=q,
        total_doses=total_doses, unique_rams=unique_rams)

@livestock_bp.route('/add_semen', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_semen():
    s = SemenInventory(
        ram_name=request.form.get('ram_name'),
        ram_id=request.form.get('ram_id', type=int) or None,
        breed=request.form.get('breed'),
        collection_date=parse_smart_date(request.form.get('collection_date')),
        quantity_doses=request.form.get('quantity_doses', type=int) or 0,
        price_per_dose=normalize_amount_to_toman(request.form.get('price_per_dose')),
        storage_location=request.form.get('storage_location'),
        expiry_date=parse_smart_date(request.form.get('expiry_date')),
        notes=request.form.get('notes')
    )
    db.session.add(s)
    db.session.commit()
    log_audit(f"ثبت اسپرم {s.ram_name} ({s.quantity_doses} دوز)")
    flash('اسپرم/سیمن ثبت شد.', 'success')
    return redirect(url_for('livestock.semen_inventory'))

@livestock_bp.route('/delete_semen/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def delete_semen(id):
    s = SemenInventory.query.get_or_404(id)
    db.session.delete(s)
    db.session.commit()
    flash("اسپرم حذف شد.", "success")
    return redirect(url_for('livestock.semen_inventory'))

@livestock_bp.route('/edit_semen/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def edit_semen(id):
    s = SemenInventory.query.get_or_404(id)
    s.ram_name = request.form.get('ram_name')
    s.ram_id = request.form.get('ram_id', type=int) or None
    s.breed = request.form.get('breed')
    s.collection_date = parse_smart_date(request.form.get('collection_date'))
    s.quantity_doses = request.form.get('quantity_doses', type=int) or 0
    s.price_per_dose = normalize_amount_to_toman(request.form.get('price_per_dose'))
    s.storage_location = request.form.get('storage_location')
    s.expiry_date = parse_smart_date(request.form.get('expiry_date'))
    s.notes = request.form.get('notes')
    db.session.commit()
    flash('اسپرم ویرایش شد.', 'success')
    return redirect(url_for('livestock.semen_inventory'))

# ============================================================
# خروجی Excel
# ============================================================

@livestock_bp.route('/export_semen_excel')
@login_required
@permission_required('can_view_livestock')
def export_semen_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'اسپرم'
    headers = ['نام قوچ', 'نژاد', 'تاریخ جمع‌آوری', 'دوز موجود', 'قیمت هر دوز', 'محل نگهداری', 'انقضا', 'توضیحات']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    doses = SemenInventory.query.order_by(SemenInventory.expiry_date).all()
    for d in doses:
        ws.append([d.ram_name, d.breed or '', str(d.collection_date or ''), d.quantity_doses,
            str(d.price_per_dose or ''), d.storage_location or '', str(d.expiry_date or ''), d.notes or ''])
    from flask import send_file
    path = os.path.join(tempfile.gettempdir(), 'semen_inventory.xlsx')
    wb.save(path)
    return send_file(path, as_attachment=True, download_name='semen_inventory.xlsx')

@livestock_bp.route('/export_drug_excel')
@login_required
@permission_required('can_view_livestock')
def export_drug_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'دارو'
    headers = ['نام دارو', 'دسته', 'موجودی', 'واحد', 'قیمت واحد', 'تأمین‌کننده', 'انقضا', 'حداقل هشدار', 'توضیحات']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    drugs = DrugInventory.query.order_by(DrugInventory.expiry_date).all()
    for d in drugs:
        ws.append([d.name, d.category, float(d.stock_quantity), d.unit, float(d.price_per_unit) if d.price_per_unit else '',
            d.supplier or '', str(d.expiry_date or ''), float(d.min_stock_alert) if d.min_stock_alert else '', d.notes or ''])
    from flask import send_file
    path = os.path.join(tempfile.gettempdir(), 'drug_inventory.xlsx')
    wb.save(path)
    return send_file(path, as_attachment=True, download_name='drug_inventory.xlsx')

# ============================================================
# انبار دارو
# ============================================================

@livestock_bp.route('/drug_inventory')
@login_required
@permission_required('can_view_livestock')
def drug_inventory():
    q = request.args.get('q', '').strip()
    filter_cat = request.args.get('category', '').strip()
    drugs = DrugInventory.query
    if q:
        drugs = drugs.filter(DrugInventory.name.ilike(f'%{q}%') | DrugInventory.supplier.ilike(f'%{q}%'))
    if filter_cat:
        drugs = drugs.filter(DrugInventory.category == filter_cat)
    drugs = drugs.order_by(DrugInventory.expiry_date).all()
    low_stock = [d for d in drugs if d.stock_quantity <= d.min_stock_alert and d.min_stock_alert > 0]
    near_expiry = [d for d in drugs if d.expiry_date and (d.expiry_date - datetime.now(UTC).date()).days <= 30]
    categories = db.session.query(DrugInventory.category).distinct().all()
    categories = [c[0] for c in categories if c[0]]
    total_value = sum(float(d.stock_quantity) * float(d.price_per_unit) for d in drugs if d.price_per_unit)
    return render_template('livestock/drug_inventory.html', drugs=drugs, low_stock=low_stock, near_expiry=near_expiry,
        q=q, filter_cat=filter_cat, categories=categories, total_value=total_value)

@livestock_bp.route('/add_drug', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_drug():
    d = DrugInventory(
        name=request.form.get('name'),
        category=request.form.get('category', 'عمومی'),
        stock_quantity=normalize_amount_to_toman(request.form.get('stock_quantity')),
        unit=request.form.get('unit', 'عدد'),
        price_per_unit=normalize_amount_to_toman(request.form.get('price_per_unit')),
        supplier=request.form.get('supplier'),
        expiry_date=parse_smart_date(request.form.get('expiry_date')),
        min_stock_alert=normalize_amount_to_toman(request.form.get('min_stock_alert')),
        notes=request.form.get('notes')
    )
    db.session.add(d)
    db.session.commit()
    log_audit(f"ثبت داروی جدید: {d.name}")
    flash(f'داروی {d.name} ثبت شد.', 'success')
    return redirect(url_for('livestock.drug_inventory'))

@livestock_bp.route('/edit_drug/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def edit_drug(id):
    d = DrugInventory.query.get_or_404(id)
    d.name = request.form.get('name')
    d.category = request.form.get('category', 'عمومی')
    d.stock_quantity = normalize_amount_to_toman(request.form.get('stock_quantity'))
    d.unit = request.form.get('unit', 'عدد')
    d.price_per_unit = normalize_amount_to_toman(request.form.get('price_per_unit'))
    d.supplier = request.form.get('supplier')
    d.expiry_date = parse_smart_date(request.form.get('expiry_date'))
    d.min_stock_alert = normalize_amount_to_toman(request.form.get('min_stock_alert'))
    d.notes = request.form.get('notes')
    db.session.commit()
    flash('دارو ویرایش شد.', 'success')
    return redirect(url_for('livestock.drug_inventory'))

@livestock_bp.route('/delete_drug/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def delete_drug(id):
    d = DrugInventory.query.get_or_404(id)
    db.session.delete(d)
    db.session.commit()
    flash('دارو حذف شد.', 'success')
    return redirect(url_for('livestock.drug_inventory'))

# ============================================================
# قرنطینه
# ============================================================

@livestock_bp.route('/add_quarantine/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def add_quarantine(id):
    sheep = Sheep.query.get_or_404(id)
    start_date = parse_smart_date(request.form.get('start_date'))
    expected_days = request.form.get('expected_days', type=int) or 14
    expected_end_date = start_date + timedelta(days=expected_days) if start_date else None
    q = QuarantineRecord(
        sheep_id=id, start_date=start_date,
        reason=request.form.get('reason', 'بیماری'),
        expected_end_date=expected_end_date,
        notes=request.form.get('notes')
    )
    db.session.add(q)
    sheep.status = 'قرنطینه'
    db.session.commit()
    log_audit(f"قرنطینه {sheep.ear_tag}: {q.reason}")
    flash(f'{sheep.ear_tag} به قرنطینه رفت.', 'warning')
    return redirect(url_for('livestock.profile', id=id))

@livestock_bp.route('/release_quarantine/<int:id>', methods=['POST'])
@login_required
@permission_required('can_view_livestock')
def release_quarantine(id):
    q = QuarantineRecord.query.get_or_404(id)
    q.end_date = datetime.now(UTC).date()
    q.is_active = False
    sheep = Sheep.query.get(q.sheep_id)
    if sheep:
        sheep.status = 'زنده و سالم'
    db.session.commit()
    log_audit(f"خروج از قرنطینه {sheep.ear_tag if sheep else 'نامشخص'}")
    flash('دام از قرنطینه خارج شد.', 'success')
    return redirect(url_for('livestock.profile', id=q.sheep_id))

# ============================================================
# اپیدمیولوژی و هشدار طغیان
# ============================================================

@livestock_bp.route('/epidemiology')
@login_required
@permission_required('can_view_livestock')
def epidemiology():
    import jdatetime
    now_j = jdatetime.datetime.now()
    # آنالیز ۳ ماه اخیر
    three_months_ago = datetime.now(UTC).date() - timedelta(days=90)
    records = MedicalRecord.query.filter(MedicalRecord.record_date >= three_months_ago).all()

    # طغیان بر اساس بهاربند: اگر > 20٪ دام‌های یه بهاربند توی ۲ هفته بیماری مشابه داشته باشند
    pens = Pen.query.all()
    alerts = []
    for pen in pens:
        pen_sheep = Sheep.query.filter_by(pen_id=pen.id, is_deleted=False).count()
        if pen_sheep < 2: continue
        two_weeks_ago = datetime.now(UTC).date() - timedelta(days=14)
        sick_pen = db.session.query(MedicalRecord.sheep_id).filter(
            MedicalRecord.record_date >= two_weeks_ago,
            MedicalRecord.sheep_id.in_(
                db.session.query(Sheep.id).filter(Sheep.pen_id == pen.id, Sheep.is_deleted == False)
            )
        ).distinct().count()
        ratio = sick_pen / pen_sheep * 100
        if ratio >= 20:
            alerts.append({'pen': pen, 'ratio': ratio, 'sick': sick_pen, 'total': pen_sheep})

    # بیماری‌های فصلی
    monthly_disease = db.session.query(
        MedicalRecord.medicine_name, func.count(MedicalRecord.id).label('cnt')
    ).filter(MedicalRecord.record_date >= three_months_ago).group_by(MedicalRecord.medicine_name).order_by(func.count(MedicalRecord.id).desc()).limit(10).all()

    # تلفات ماهانه برای نمودار (۶ ماه)
    monthly_deaths = []
    monthly_labels = []
    for i in range(5, -1, -1):
        m = now_j.month - i
        y = now_j.year
        while m < 1:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        j_start = jdatetime.date(y, m, 1)
        if m == 12:
            j_end = jdatetime.date(y + 1, 1, 1)
        else:
            j_end = jdatetime.date(y, m + 1, 1)
        g_start = j_start.togregorian()
        g_end = j_end.togregorian()
        deaths_month = Sheep.query.filter(
            Sheep.is_deleted == False, Sheep.status == 'تلف شده',
            Sheep.entry_date >= g_start, Sheep.entry_date < g_end
        ).count()
        monthly_deaths.append(deaths_month)
        month_names = {1:'فروردین',2:'اردیبهشت',3:'خرداد',4:'تیر',5:'مرداد',6:'شهریور',7:'مهر',8:'آبان',9:'آذر',10:'دی',11:'بهمن',12:'اسفند'}
        monthly_labels.append(month_names[m])

    # تلفات
    deaths = Sheep.query.filter(
        Sheep.is_deleted == False,
        Sheep.status == 'تلف شده',
        Sheep.entry_date >= three_months_ago
    ).count()

    return render_template('livestock/epidemiology.html', alerts=alerts,
        monthly_disease=monthly_disease, deaths=deaths, records_count=len(records),
        monthly_deaths=monthly_deaths, monthly_labels=monthly_labels)

# ============================================================
# تقویم فحلی پیش‌بینی شده + آلرت
# ============================================================

@livestock_bp.route('/heat_alerts')
@login_required
@permission_required('can_view_livestock')
def heat_alerts():
    import jdatetime
    now = datetime.now(UTC).date()
    # میش‌هایی که last_heat_date دارند و آبستن نیستند
    eligible = Sheep.query.filter(
        Sheep.is_deleted == False,
        Sheep.gender.in_(['میش', 'بره ماده']),
        Sheep.last_heat_date != None,
        Sheep.status.notin_(['آبستن', 'قرنطینه', 'تلف شده', 'فروخته شده'])
    ).all()

    heat_list = []
    for s in eligible:
        # میانگین سیکل از تاریخ‌های جفت‌اندازی قبلی
        matings = MatingRecord.query.filter_by(sheep_id=s.id).order_by(MatingRecord.mating_date.desc()).limit(3).all()
        if len(matings) >= 2:
            # محاسبه میانگین فاصله بین جفت‌اندازی‌ها
            intervals = []
            for i in range(len(matings) - 1):
                diff = (matings[i].mating_date - matings[i + 1].mating_date).days
                if 14 <= diff <= 21:
                    intervals.append(diff)
            avg_cycle = sum(intervals) / len(intervals) if intervals else 17
        else:
            avg_cycle = 17  # پیش‌فرض استاندارد

        next_heat = s.last_heat_date + timedelta(days=int(avg_cycle))
        days_until = (next_heat - now).days
        if -3 <= days_until <= 5:  # بازه هشدار
            heat_list.append({
                'sheep': s, 'next_heat': next_heat,
                'days_until': days_until, 'cycle': int(avg_cycle),
                'confidence': 'بالا' if len(matings) >= 2 else 'متوسط'
            })

    heat_list.sort(key=lambda x: x['days_until'])
    rams = Sheep.query.filter(Sheep.is_deleted == False, Sheep.gender == 'قوچ').order_by(Sheep.ear_tag).all()
    semen_doses = SemenInventory.query.filter(SemenInventory.quantity_doses > 0).order_by(SemenInventory.ram_name).all()
    return render_template('livestock/heat_alerts.html', heat_list=heat_list, rams=rams, semen_doses=semen_doses)

# ============================================================
# شاخص ژنتیکی ترکیبی (Genetic Index)
# ============================================================

@livestock_bp.route('/genetic_report/<int:id>')
@login_required
@permission_required('can_view_livestock')
def genetic_report(id):
    sheep = Sheep.query.get_or_404(id)
    if sheep.gender not in ('قوچ', 'میش'):
        flash('فقط برای قوچ و میش شاخص ژنتیکی قابل محاسبه است.', 'warning')
        return redirect(url_for('livestock.profile', id=id))

    # محاسبه شاخص ترکیبی
    # وزن تولد (از اولین WeightRecord نزدیک به تولد)
    birth_weight = WeightRecord.query.filter_by(sheep_id=id).order_by(WeightRecord.record_date).first()

    # تعداد زایش (برای میش) / تعداد نتاج (برای قوچ)
    if sheep.gender == 'میش':
        births = BirthRecord.query.filter_by(mother_id=id).all()
        total_births = len(births)
        total_lambs = sum(b.lambs_count for b in births)
        live_births = sum(1 for b in births if b.status == 'موفق')
        twin_rate = sum(1 for b in births if b.lambs_count >= 2) / total_births if total_births else 0
        # شاخص: (نرخ چندقلوزایی × ۳۵) + (نرخ زایش موفق × ۳۰) + (تعداد کل بره × ۲۰) + (وزن تولد شاخص × ۱۵)
        index_score = (twin_rate * 35) + ((live_births / total_births if total_births else 0) * 30) + (min(total_lambs, 20) * 2) + (float(birth_weight.weight or 0) * 1.5 if birth_weight else 7.5)
    else:  # قوچ
        offspring = Sheep.query.filter(db.or_(Sheep.father_id == id)).count()
        father_births = BirthRecord.query.filter_by(father_id=id).all()
        fb_count = len(father_births)
        twin_fb = sum(1 for b in father_births if b.lambs_count >= 2)
        twin_rate = twin_fb / fb_count if fb_count else 0
        index_score = (offspring * 3) + (twin_rate * 40) + (float(birth_weight.weight or 0) * 1.5 if birth_weight else 7.5)

    # بررسی هم‌خونی
    inbreeding = False
    inbreeding_pct = 0
    if sheep.mother_id and sheep.father_id:
        common = db.session.query(Sheep).filter(
            Sheep.id.in_([sheep.mother_id, sheep.father_id])
        ).first()
        # ساده: اگر مادر و پدر هر دو یک parent مشترک داشته باشند
        mother = Sheep.query.get(sheep.mother_id)
        father = Sheep.query.get(sheep.father_id)
        if mother and father:
            mother_parents = {mother.mother_id, mother.father_id}
            father_parents = {father.mother_id, father.father_id}
            common_parents = mother_parents & father_parents
            common_parents.discard(None)
            if common_parents:
                inbreeding = True
                inbreeding_pct = 25  # هم‌خونی درجه یک
            # بررسی نسل دوم
            for mp in mother_parents:
                if mp:
                    mp_sheep = Sheep.query.get(mp)
                    if mp_sheep:
                        for fp in father_parents:
                            if fp:
                                fp_sheep = Sheep.query.get(fp)
                                if fp_sheep and (mp_sheep.mother_id == fp_sheep.mother_id or mp_sheep.father_id == fp_sheep.father_id):
                                    inbreeding = True
                                    inbreeding_pct = 12.5
                                    break

    # رتبه در گله
    all_same_gender = Sheep.query.filter(Sheep.gender == sheep.gender, Sheep.is_deleted == False).count()

    return render_template('livestock/genetic_report.html', sheep=sheep,
        index_score=min(index_score, 100), inbreeding=inbreeding,
        inbreeding_pct=inbreeding_pct, total_gender=all_same_gender)

# ============================================================
# گواهی سلامت دام (PDF/چاپ)
# ============================================================

@livestock_bp.route('/health_certificate/<int:id>')
@login_required
@permission_required('can_view_livestock')
def health_certificate(id):
    sheep = Sheep.query.get_or_404(id)
    today = datetime.now(UTC).date()
    treatments = MedicalRecord.query.filter_by(sheep_id=id).order_by(MedicalRecord.record_date.desc()).limit(20).all()
    vaccines = [m for m in treatments if m.action_type == 'واکسن']
    weights = WeightRecord.query.filter_by(sheep_id=id).order_by(WeightRecord.record_date.desc()).limit(5).all()
    import jdatetime
    age_months = ((today - sheep.birth_date).days // 30) if sheep.birth_date else 0
    return render_template('livestock/health_certificate.html', sheep=sheep, vaccines=vaccines, weights=weights, today=today, age_months=age_months)

# ============================================================
# داشبورد دامپزشکی با KPI
# ============================================================

@livestock_bp.route('/vet_dashboard')
@login_required
@permission_required('can_view_livestock')
def vet_dashboard():
    today = datetime.now(UTC).date()
    thirty_days_ago = today - timedelta(days=30)

    total_sheep = Sheep.query.filter(Sheep.is_deleted == False).count()
    sick_count = Sheep.query.filter(Sheep.status == 'بیمار', Sheep.is_deleted == False).count()
    under_treatment = Sheep.query.filter(Sheep.status == 'تحت درمان', Sheep.is_deleted == False).count()
    pregnant_count = Sheep.query.filter(Sheep.status == 'آبستن', Sheep.is_deleted == False).count()
    quarantined = Sheep.query.filter(Sheep.status == 'قرنطینه', Sheep.is_deleted == False).count()
    deaths_30d = Sheep.query.filter(Sheep.status == 'تلف شده', Sheep.entry_date >= thirty_days_ago, Sheep.is_deleted == False).count()

    # درمان‌های ۳۰ روز اخیر
    treatments_30d = MedicalRecord.query.filter(MedicalRecord.record_date >= thirty_days_ago).count()

    # داروهای در شرف انقضا (۳۰ روز آینده)
    expiring_drugs = DrugInventory.query.filter(
        DrugInventory.expiry_date != None,
        DrugInventory.expiry_date <= today + timedelta(days=30),
        DrugInventory.expiry_date >= today
    ).count()

    # داروهای کم‌موجودی
    low_stock_drugs = DrugInventory.query.filter(
        DrugInventory.stock_quantity <= DrugInventory.min_stock_alert,
        DrugInventory.min_stock_alert > 0
    ).count()

    # جفت‌اندازی‌های ۳۰ روز اخیر
    matings_30d = MatingRecord.query.filter(MatingRecord.mating_date >= thirty_days_ago).count()

    # زایش‌های ۳۰ روز اخیر
    births_30d = BirthRecord.query.filter(BirthRecord.birth_date >= thirty_days_ago).count()

    kpi = {
        'total': total_sheep, 'sick': sick_count, 'under_treatment': under_treatment,
        'pregnant': pregnant_count, 'quarantined': quarantined, 'deaths': deaths_30d,
        'treatments': treatments_30d, 'expiring_drugs': expiring_drugs,
        'low_stock_drugs': low_stock_drugs, 'matings': matings_30d, 'births': births_30d
    }

    # آلرت‌ها
    alerts = []
    if sick_count > 0: alerts.append(f"{sick_count} دام بیمار نیازمند درمان")
    if quarantined > 0: alerts.append(f"{quarantined} دام در قرنطینه")
    if expiring_drugs > 0: alerts.append(f"{expiring_drugs} داروی در شرف انقضا")
    if low_stock_drugs > 0: alerts.append(f"{low_stock_drugs} داروی کم‌موجودی")
    mortality_rate = (deaths_30d / total_sheep * 100) if total_sheep > 0 else 0
    if mortality_rate > 2: alerts.append(f"نرخ تلفات {mortality_rate:.1f}٪ در ۳۰ روز اخیر (هشدار)")

    # فعالیت‌های اخیر
    recent_treatments = MedicalRecord.query.options(db.joinedload(MedicalRecord.drug)).order_by(MedicalRecord.id.desc()).limit(10).all()
    recent_matings = MatingRecord.query.order_by(MatingRecord.id.desc()).limit(5).all()

    return render_template('livestock/vet_dashboard.html', kpi=kpi, alerts=alerts, mortality_rate=mortality_rate,
        recent_treatments=recent_treatments, recent_matings=recent_matings)